from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from cerros.models.mountains import Mountain, MountainPrefix
from cerros.models.nomenclatura import NomenclaturaSummit, IGMMap
from cerros.models.routes import Route
from cerros.models.andinists import Club, Andinist
from cerros.models.geography import Country, Region, MountainGroup
from cerros.models.ascents import Ascent
from cerros.models.references import Reference
import datetime

from django.db.models import Value as V
from django.db.models.functions import Concat   


class Command(BaseCommand):
  help = 'Populate the database with the Cerros data'

  def handle(self, *args, **options):
    # CARGAR COUNTRIES, CLUBS, REGIONS, MOUNTAIN GROUPS PRIMERO???

    # IGM MAPS VA PRIMERO
    wb = load_workbook(filename = 'IGM Rectangles.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
      id, name, file_id = row
      igm_map = IGMMap()
      igm_map.name = name
      igm_map.file_id = file_id
      igm_map.save()

    # NOMENCLATURA VA SEGUNDO, PERO HAY QUE CORREGIRLE EL EXCEL
    wb = load_workbook(filename = 'NomenclaturaObjects.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=11, values_only=True):
      id, nom_id, revision_code, name, altitude_igm, latitude_south, longitude_west, observations, comment, igm_rectangle_id, mountain_id = row
      nomenclatura_summit = NomenclaturaSummit()
      nomenclatura_summit.id_nomenclatura = id
      nomenclatura_summit.cod_revision = revision_code
      nomenclatura_summit.name = name
      nomenclatura_summit.altitude_igm = altitude_igm
      nomenclatura_summit.latitude = latitude_south
      nomenclatura_summit.longitude = longitude_west
      nomenclatura_summit.observations = observations
      nomenclatura_summit.comment = comment
      nomenclatura_summit.igm_rectangle = IGMMap.objects.get(id=igm_rectangle_id)
      nomenclatura_summit.save()

    # DESPUÉS VA MOUNTAINS
    wb = load_workbook(filename = 'Mountains.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=17, values_only=True):
      if row[0] is None:
        continue
      id, name, prefix, altitude, nom_id, alternative_name_prefix, alternative_name, secondary_altitude, secondary_altitude_name, previously_ascended, latitude, longitude, ancestry, img_url, img_author, unregistered_non_sport_ascent, unregistered_sport_ascent = row
      mountain = Mountain()
      if MountainPrefix.objects.filter(prefix=prefix).exists():
        mountain.prefix = MountainPrefix.objects.get(prefix=prefix)
      else:
        mountain_prefix = MountainPrefix()
        mountain_prefix.prefix = prefix
        mountain_prefix.save()
        mountain.prefix = mountain_prefix
      mountain.name = name
      mountain.latitude = latitude
      mountain.longitude = longitude
      mountain.altitude = altitude
      mountain.altitude_igm = altitude
      if unregistered_non_sport_ascent == 'true':
        mountain.unregistered_non_sport_ascent = True
      if unregistered_sport_ascent == 'true':
        mountain.unregistered_sport_ascent = True
      mountain.main_altitude_source = Mountain.IGM
      if ancestry:
        if ancestry < id:
          parent = Mountain.objects.get(id=ancestry)
          mountain.parent_mountain = parent
      if nom_id:
        nomenclatura_mountain = NomenclaturaSummit.objects.get(id_nomenclatura=nom_id)
        mountain.nomenclatura_mountain = nomenclatura_mountain
      mountain.save()

    # DESPUÉS VA ROUTES
    wb = load_workbook(filename = 'Routes.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
      id, name, mountain, unregistered_non_sport_ascent, unregistered_sport_ascent = row
      route = Route()
      route.name = name
      route.mountain = Mountain.objects.get(id=mountain)
      if unregistered_non_sport_ascent == 'true':
        route.unregistered_non_sport_ascent = True
      if unregistered_sport_ascent == 'true':
        route.unregistered_sport_ascent = True
      route.save()

    # DESPUÉS VA ANDINISTS
    wb = load_workbook(filename = 'Andinists.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
      id, name, surname, country_to_s, club_to_s, gender = row
      andinist = Andinist()
      andinist.name = name
      andinist.surname = surname
      
      if gender:
        andinist.gender = gender
      andinist.save()

      if country_to_s:
        if Country.objects.filter(name=country_to_s).exists():
          andinist.nationalities.add(Country.objects.get(name=country_to_s))
        else:
          country = Country()
          country.name = country_to_s
          country.save()
          andinist.nationalities.add(country)
      if club_to_s:
        if Club.objects.filter(name=club_to_s).exists():
          andinist.clubs.add(Club.objects.get(name=club_to_s))
        else:
          club = Club()
          club.name = club_to_s
          club.save()
          andinist.clubs.add(club)

        andinist.save()

    # DESPUES VA ASCENTS
    wb = load_workbook(filename = 'Ascents.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
      if row[0] is None:
        continue
      id, year, month, day, route, name, first_absolute, first_ascent, andinists_to_s = row
      ascent = Ascent()
      if year:
        if month:
          if day:
            ascent.date = datetime.date(year, month, day)
            ascent.date_format = Ascent.DAYMONTHYEAR
          else:
            ascent.date = datetime.date(year, month, 1)
            ascent.date_format = Ascent.MONTHYEAR
        else:
          ascent.date = datetime.date(year, 1, 1)
          ascent.date_format = Ascent.YEAR
      ascent.route = Route.objects.get(id=route)
      ascent.name = name
      ascent.save()
      if andinists_to_s:
        andinists = andinists_to_s.split(', ')
        for andinist in andinists:
          andinist_obj = Andinist.objects.annotate(full_name=Concat('name', V(' '), 'surname')).filter(full_name=andinist).first()
          ascent.andinists.add(andinist_obj)
        ascent.save()

    for route in Route.objects.all():
      route.save()

    for mountain in Mountain.objects.all():
      mountain.save()
      
    for andinist in Andinist.objects.all():
      andinist.save()

    # REFERENCES
    wb = load_workbook(filename = 'Sources.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
      id, name, link, referenceable_type, referenceable_id = row
      reference = Reference()
      reference.title = name
      reference.url = link
      reference.save()
      if referenceable_type == 'Mountain':
        mtn = Mountain.objects.get(id=referenceable_id)
        mtn.references.add(reference)
        mtn.save()
      elif referenceable_type == 'Route':
        route = Route.objects.get(id=referenceable_id)
        route.references.add(reference)
        route.save()
      elif referenceable_type == 'Ascent':
        ascent = Ascent.objects.get(id=referenceable_id)
        ascent.references.add(reference)
        ascent.save()



